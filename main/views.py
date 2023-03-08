from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView, LogoutView
from django.core.exceptions import PermissionDenied
from django.urls import reverse_lazy
from .views_student import get_students_manager, get_students_teacher, get_paid_groups
from .forms import *
import openpyxl as xl
from django.shortcuts import render


class MainLoginView(LoginView):
    template_name = 'login.html'
    form_class = AuthUserForm
    success_url = reverse_lazy('/')


class MainLogout(LogoutView):
    next_page = '/login'


def error404_handler(request, exception):
    error = "Ошибка 404"
    msg = "Страницы с такой ссылкой не существует"
    response = render(request, "page_error.html", {
                      'error_msg': msg, 'error_title': error})
    response.status_code = 404
    return response


def error403_handler(request, exception):
    error = "Ошибка 403"
    msg = "У вас нет прав для совершения этого действия"
    response = render(request, "page_error.html", {
                      'error_msg': msg, 'error_title': error})
    response.status_code = 404
    return response


@login_required(login_url='/login')
def home(request):
    if not request.user.has_perm('main.can_edit'):
        user_can_edit = False
    else:
        user_can_edit = True

    user = request.user
    directions = Direction.objects.all()
    users = User.objects.all()
    if not user.userprofile.is_teacher:
        groups = Group.objects.all()
        paid_groups = PaidGroup.objects.all()
    else:
        groups = Group.objects.filter(teacher=user)
        paid_groups = PaidGroup.objects.filter(teacher=user)

    if not user.userprofile.is_teacher:
        students = get_students_manager(request)
    else:
        students = get_students_teacher(request)

    students = get_paid_groups(students, 0)

    return render(request, 'home.html', {'students': students,
                                         'user_can_edit': user_can_edit,
                                         'directions': directions,
                                         'teachers': users, 'user': user,
                                         'groups': groups, 'paid_groups': paid_groups})


def get_xlsx(file):
    wb = xl.load_workbook(filename=file)
    worksheet = wb.active
    for i in range(2, worksheet.max_row):
        student = Student()
        student.petition_date = worksheet.cell(i, 3).value
        last_name = str(worksheet.cell(i, 6).value).capitalize()
        if last_name.upper().count('РЕЗЕРВ') > 0:
            last_name = last_name.upper().replace('РЕЗЕРВ', '').capitalize()
        student.last_name = last_name
        student.first_name = worksheet.cell(i, 7).value
        student.middle_name = worksheet.cell(i, 8).value
        student.birthday = worksheet.cell(i, 9).value
        print(i)
        student.parent_name = str(worksheet.cell(i, 12).value).capitalize() + " " + \
            str(worksheet.cell(i, 13).value).capitalize() + " " + \
            str(worksheet.cell(i, 14).value).capitalize()

        student.parent_number = worksheet.cell(i, 15).value
        student.email = worksheet.cell(i, 16).value
        for col in range(17, 22):
            if worksheet.cell(i, col).value is not None:
                if col == 17:
                    student.social_category = 0
                if col == 18:
                    student.social_category = 2
                if col == 19:
                    student.social_category = 1
                if col == 20:
                    student.social_category = 3
                if col == 21:
                    student.social_category = 4
        student.status = 1
        student.school = worksheet.cell(i, 11).value
        student.comment = worksheet.cell(i, 26).value
        student.direction = Direction.objects.get(
            direction_name=worksheet.cell(i, 24).value)
        if Student.objects.filter(last_name=student.last_name,
                                  first_name=student.first_name,
                                  middle_name=student.middle_name).count() == 0:
            student.save()
        else:
            student_old = Student.objects.filter(last_name=student.last_name,
                                                 first_name=student.first_name,
                                                 middle_name=student.middle_name).first()
            if student_old:
                student_old.comment = str(
                    student.comment) + '\n' + str(student.direction.direction_name)
                student_old.save()


@login_required(login_url='/login')
def load_file(request):
    if not request.user.has_perm('main.can_edit'):
        raise PermissionDenied
    if request.POST:
        load_form = LoadTableForm(request.POST, request.FILES)
        if load_form.is_valid():
            load_form.save()

            filename = LoadTable.objects.last()
            get_xlsx(filename.file_name)
    else:
        load_form = LoadTableForm
    return render(request, 'load.html', {'load_form': load_form})
