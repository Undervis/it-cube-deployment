from rest_framework.viewsets import ModelViewSet
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import permissions
from .serializers import StudentsSerializer
from .models import *
import openpyxl as xl
from it_cube.settings import BASE_DIR
from django.shortcuts import redirect

class Table:
    def __init__(self):
        self.month: str
        self.dates = []
        self.students = []
        self.themes = []
        self.tb = []

class HomeView(ModelViewSet):
    students = Student.objects.all()
    queryset = students
    serializer_class = StudentsSerializer

    def get_queryset(self):
        request = self.request.query_params
        if request.get('direction'):
            queryset = Student.objects.filter(direction=Direction.objects.get(id=request.get("direction")))
        elif request.get('group'):
            queryset = Student.objects.filter(group=Group.objects.get(id=request.get("group")))
        elif request.get('paidgroup'):
            queryset = Student.objects.filter(paid_group=PaidGroup.objects.get(id=request.get("paidgroup")))
        else:
            queryset = Student.objects.all()
        return queryset
    
class JournalView(APIView):
    renderer_classes = [JSONRenderer]
    permission_classes = (permissions.AllowAny,)

    def get(self, request, month):
        group_name: str
        user = request.user

        # Проверка авторизации пользователя и перевод на логин если пользователь не авторизован
        if not user.is_authenticated:
            return redirect("/login")
        # Проверка наличия групп
        if Group.objects.all().count() or PaidGroup.objects.all().count():
            # Если пользователь учитель то ему выдаются только его группы, иначе выдаются все группы
            if not user.userprofile.is_teacher:
                groups = Group.objects.all()
                paid_groups = PaidGroup.objects.all()
            else:
                groups = Group.objects.filter(teacher=user)
                paid_groups = PaidGroup.objects.filter(teacher=user)

            # Проверка параметров GET, выдача группы по заданному параметру
            if not request.GET:
                first_group = groups.union(paid_groups).first()
                group_name = first_group.group_name
                students = Student.objects.filter(group=first_group)
                redirect_url = f"/api/journal/{month}/"
            if request.GET.get('group'):
                group_name = Group.objects.get(id=request.GET.get('group')).group_name
                students = Student.objects.filter(group__id=request.GET.get('group'))
                redirect_url = f"/api/journal/{month}/?group={request.GET.get('group')}"
            if request.GET.get('paidgroup'):
                group_name = PaidGroup.objects.get(id=request.GET.get('paidgroup')).group_name
                students = Student.objects.filter(paid_group__id=request.GET.get('paidgroup'))
                redirect_url = f"/api/journal/{month}/?paidgroup={request.GET.get('paidgroup')}"

            # Конвертация списка учащихся в формат JSON
            students_js = []
            for s in students:
                students_js.append({"first_name": s.first_name, "last_name": s.last_name, "status": s.status,
                                    "delete_date": s.delete_date})

            # Кол-во столбцов для дат и тем
            group_marks = {}
            group_tb = []
            cols = 9
            try:
                # Проверка названия группы на наличие символа "/"
                if group_name.count("/") > 0:
                    group_name = group_name.replace("/", ",")
                # Подгрузка файла EXCEL по название группы
                file_name = str(BASE_DIR) + '/journal/' + group_name + '.xlsx'
                try:
                    wb = xl.load_workbook(file_name)
                except Exception as e:
                    print("Ошибка при загрузке файла", e)
                    xl.Workbook().save(file_name)
                    wb = xl.load_workbook(file_name)
                    wb.create_sheet(month.capitalize())
                    wb.create_sheet("ТБ")
                    wb.save(file_name)
                    
                    return redirect(redirect_url)
                # Составление ответа из таблицы по выбраному месяцу
                try:
                    sheet = wb.get_sheet_by_name(month.capitalize())
                    group_marks = {"month": str, "dates": [], "students_marks": [], "themes": []}
                    group_marks["month"] = month.capitalize()
                    for col in range(2, cols + 2):
                        if sheet.cell(row=1, column=col).value:
                            group_marks["dates"].append(str(sheet.cell(row=1, column=col).value))
                        else:
                            group_marks["dates"].append("")
                    row = 2
                    while True:
                        if sheet.cell(row=row, column=1).value:
                            group_marks["students_marks"].append({"row": row - 2, "marks": []})
                            for col in range(2, cols + 2):
                                if sheet.cell(row=row, column=col).value == "НБ":
                                    group_marks["students_marks"][row - 2]["marks"].append(1)
                                elif sheet.cell(row=row, column=col).value == "Б":
                                    group_marks["students_marks"][row - 2]["marks"].append(2)
                                else:
                                    group_marks["students_marks"][row - 2]["marks"].append(0)
                            row += 1
                        else:
                            try:
                                for row in range(12 + 1):
                                    group_marks["students_marks"].append({"row": row, "marks": []})
                                    for col in range(cols + 2):
                                        group_marks["students_marks"][row]["marks"].append(0)
                                break
                            except Exception as e:
                                print(e)
                    try:
                        for row in range(2, cols + 2):
                            theme_date = sheet.cell(row=row, column=18).value
                            theme_time = sheet.cell(row=row, column=19).value
                            theme_title = sheet.cell(row=row, column=20).value
                            group_marks["themes"].append({"date": str(theme_date) if theme_date else "",
                                                "time": str(theme_time) if theme_time else "",
                                                "theme": str(theme_title) if theme_title else ""})
                    except Exception as e:
                        print(e)
                        for row in range(2, cols + 2):
                            group_marks["themes"].append({"date": "",
                                                "time": "",
                                                "theme": ""})
                        
                except Exception as e:
                    print("Ошибка в таблице отметок:", e)
                    wb = xl.load_workbook(file_name)
                    wb.create_sheet(month.capitalize())
                    wb.save(file_name)

                    return redirect(redirect_url)
                
                try:
                    sheet_tb = wb.get_sheet_by_name('ТБ')
                    group_tb = []

                    for row in range(1, 16 + 1):
                        date_bs = sheet_tb.cell(row=row + 1, column=2).value
                        theme_bs = sheet_tb.cell(row=row + 1, column=3).value
                        date_pdd = sheet_tb.cell(row=row + 1, column=5).value
                        theme_pdd = sheet_tb.cell(row=row + 1, column=6).value
                        group_tb.append({
                            "date_bs": date_bs if date_bs else "",
                            "theme_bs": theme_bs if theme_bs else "",
                            "date_pdd": date_pdd if date_pdd else "",
                            "theme_pdd": theme_pdd if theme_pdd else ""
                        })

                except Exception as e:
                    print("Ошибка в таблице техники безопасности:", e)
                    wb = xl.load_workbook(file_name)
                    wb.create_sheet("ТБ")
                    wb.save(file_name)

                    return redirect(redirect_url)
                wb.close()
            except Exception as e:
                print("Ошибка в общей обработке таблицы:", e)

            

            return Response({"group_marks": group_marks, "group_tb": group_tb, "students": students_js})