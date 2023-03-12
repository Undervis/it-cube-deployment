from .models import Group, PaidGroup
from .forms import *
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
import openpyxl as xl
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from it_cube.settings import BASE_DIR

@login_required(login_url='/login')
def journal(request):
    user = request.user
    if not request.user.has_perm('main.can_edit'):
        user_can_edit = False
    else:
        user_can_edit = True
    if not user.userprofile.is_teacher:
        groups = Group.objects.all()
        paid_groups = PaidGroup.objects.all()
    else:
        groups = Group.objects.filter(teacher=user)
        paid_groups = PaidGroup.objects.filter(teacher=user)
            
    return render(request, 'journal.html', {'groups': groups, 'paid_groups': paid_groups, 'user_can_edit': user_can_edit, 'user': user})


class Table:
    def __init__(self):
        self.month: str
        self.dates = []
        self.students = []
        self.themes = []
        self.tb = []

"""@login_required(login_url='/login')
def get_table(request):
    """

@csrf_exempt
@login_required(login_url='/login')
def load_json(request):
    if request.POST:
        table_data = json.loads(str(request.POST.get('data')))
        print(table_data)
        group_name = str(table_data['group'])
        if group_name.count('/') > 0:
            group_name = group_name.replace("/", ',')
        file_name = str(BASE_DIR) + '/journal/' + group_name + '.xlsx'
        try:
            wb = xl.load_workbook(file_name)
        except FileNotFoundError:
            xl.Workbook().save(file_name)
            wb = xl.load_workbook(file_name)

        try:
            sheet = wb.get_sheet_by_name(table_data['month'])
        except:
            sheet = wb.create_sheet(table_data['month'])

        sheet.cell(row=1, column=1).value = 'Фамилия Имя'
        sheet.column_dimensions['A'].width = 20
        for i in range(len(table_data['dates'])):
            sheet.cell(row=1, column=i + 2).value = table_data['dates'][i]

        for x in range(len(table_data['students'])):
            sheet.cell(row=x + 2, column=1).value = table_data['students'][x]['name']
            for y in range(len(table_data['students'][x]['marks'])):
                if table_data['students'][x]['marks'][y] == '0':
                    sheet.cell(row=x + 2, column=y + 2).value = ""
                elif table_data['students'][x]['marks'][y] == '1':
                    sheet.cell(row=x + 2, column=y + 2).value = "НБ"
                elif table_data['students'][x]['marks'][y] == '2':
                    sheet.cell(row=x + 2, column=y + 2).value = "Б"

        sheet.cell(row=1, column=18).value = 'Дата'
        sheet.cell(row=1, column=19).value = 'Часы'
        sheet.cell(row=1, column=20).value = 'Тема'
        sheet.column_dimensions['T'].width = 50
        for row in range(len(table_data['themes'])):
            sheet.cell(row=row + 2, column=18).value = table_data['themes'][row]['date']
            sheet.cell(row=row + 2, column=19).value = table_data['themes'][row]['time']
            sheet.cell(row=row + 2, column=20).value = table_data['themes'][row]['theme']

        try:
            sheet_tb = wb.get_sheet_by_name('ТБ')
        except:
            sheet_tb = wb.create_sheet('ТБ')

        sheet_tb.cell(row=1, column=1).value = "Фамилия имя"
        sheet_tb.cell(row=1, column=2).value = "Дата"
        sheet_tb.cell(row=1, column=3).value = "Тема"
        sheet_tb.cell(row=1, column=4).value = "Подпись"
        sheet_tb.cell(row=1, column=5).value = "Дата"
        sheet_tb.cell(row=1, column=6).value = "Тема"
        sheet_tb.cell(row=1, column=7).value = "Подпись"
        sheet_tb.column_dimensions['A'].width = 20
        sheet_tb.column_dimensions['C'].width = 30
        sheet_tb.column_dimensions['F'].width = 30

        for row in range(len(table_data['tb'])):
            sheet_tb.cell(row=row + 2, column=1).value = table_data['tb'][row]['name']
            sheet_tb.cell(row=row + 2, column=2).value = table_data['tb'][row]['date-bs']
            sheet_tb.cell(row=row + 2, column=3).value = table_data['tb'][row]['theme-bs']
            sheet_tb.cell(row=row + 2, column=5).value = table_data['tb'][row]['date-pdd']
            sheet_tb.cell(row=row + 2, column=6).value = table_data['tb'][row]['theme-pdd']

        wb.save(file_name)

    return JsonResponse({"msg": 'Success'})